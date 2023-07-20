import tkinter as tk
import tkinter.ttk as ttk
import tkinter.messagebox
from tkinter import filedialog
from tkintertable import TableModel, TableCanvas
from DataProcess import *
from functools import partial
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def neat_file(path: str):
    wb = load_workbook(path)
    ws = wb.active
    print(ws.columns)
    # 自动调整列宽
    for column_cells in ws.columns:
        print(column_cells)
        max_length = 0
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length * 2 + 2
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(path)


def initial_button(excel_dict):
    global dict_variable
    dict_variable = excel_dict


def format_table(tablecanvas: TableCanvas, table_dict: dict):
    max_len_name = 5
    max_len_room = 5
    for line in table_dict:
        max_len_name = max(len(table_dict[line]['公司名称']), max_len_name)
        max_len_room = max(len(table_dict[line]['房号']), max_len_room)
    tablecanvas.model.columnwidths['公司名称'] = max_len_name*20
    tablecanvas.model.columnwidths['房号'] = max_len_room*20


def choose_excel(tablecanvas: TableCanvas,  default_excel_filepath="合同表格 copy.xlsx"):
    """
    open a dialog to let the user choose the excel file they want to open

    tablecanvas: the canvas to show the result
    dict_variable: a variable to store the data dict
    """
    excel_file = filedialog.askopenfilename()
    if excel_file == "":
        excel_file = default_excel_filepath
    excel_dict = excel_to_dict(excel_file)
    global dict_variable
    # print(dict_variable.get())
    dict_variable = excel_dict
    excel_model = TableModel()
    excel_model.importDict(excel_dict)
    tablecanvas.updateModel(excel_model)
    format_table(tablecanvas, excel_dict)
    tablecanvas.redrawTable()


def query(month_combo: ttk.Combobox, pay_way_combo: ttk.Combobox):
    month = month_combo.get()
    pay_way = pay_way_combo.get()
    print("月份", month)
    print(pay_way)
    query_window = tk.Toplevel()
    query_window.title("查询界面")
    temp_result_dict = {}
    global dict_variable
    if pay_way != "所有":
        for line in dict_variable:
            if dict_variable[line]["付款方式"] == pay_way:
                temp_result_dict[line] = dict_variable[line]
    else:
        temp_result_dict = dict_variable.copy()
    result_dict = {}
    if pay_way == "季度":
        result_dict = season_pay(month, temp_result_dict)
        print("季度统计完成")
        print(result_dict)
    if pay_way == "半年":
        result_dict = halfyear_pay(month, temp_result_dict)
        print("半年统计完成")
        print(result_dict)
    if pay_way == "年度":
        result_dict = year_pay(month, temp_result_dict)
        print("年度统计完成")
        print(result_dict)
    if pay_way == "所有":
        season_dict = season_pay(month, temp_result_dict)
        halfyear_dict = halfyear_pay(month, temp_result_dict)
        year_dict = year_pay(month, temp_result_dict)
        result_dict.update(season_dict)
        print(result_dict)
        result_dict.update(halfyear_dict)
        print(result_dict)
        result_dict.update(year_dict)
        print("统计完成")
        print(result_dict)
    result_frame = tk.Frame(query_window)
    result_frame.pack()
    if not result_dict:
        result_label = tk.Label(result_frame, text="无")
        result_label.pack()
        return
    result_model = TableModel()
    result_model.importDict(result_dict)
    result_tablecanvas = TableCanvas(result_frame, model=result_model)

    format_table(result_tablecanvas, result_dict)
    result_tablecanvas.show()
    rent_info = tk.Label(query_window, text="房租总额")
    rent_info.pack()
    rent_label = tk.Label(query_window, text=sum_rent(result_dict))
    rent_label.pack()
    management_info = tk.Label(query_window, text="物管费总额")
    management_info.pack()
    management_label = tk.Label(query_window, text=sum_management(result_dict))
    management_label.pack()

    save_button = tk.Button(query_window, text="保存",
                            command=partial(save_result, result_model, month_combo, pay_way_combo))
    save_button.pack()
    # print(temp_result_dict)
    query_window.mainloop()


def modify():
    global dict_variable
    modify_window = tk.Toplevel()
    modify_window.title("修改界面")
    modify_frame = tk.Frame(modify_window)
    modify_frame.pack()
    result_model = TableModel()
    result_model.importDict(dict_variable)
    result_table = TableCanvas(modify_frame, model=result_model)
    format_table(result_table, dict_variable)
    result_table.show()
    add_row = tk.Button(modify_window, text="新增行",
                        command=partial(add_new_row, result_table))
    add_row.pack()
    delete_row = tk.Button(modify_window, text="删除行",
                           command=partial(delete_current_row, result_table))
    delete_row.pack()
    save_model = tk.Button(modify_window, text="保存",
                           command=partial(save, result_model))
    save_model.pack()


def save_result(model: TableModel, month_combox: ttk.Combobox, pay_way_combox: ttk.Combobox):
    global dict_variable
    print("saving")
    file_path = filedialog.asksaveasfilename(defaultextension='.xlsx',
                                             filetypes=[('Excel文件', '*.xlsx'), ('所有文件', '*.*')])
    excel_data = {}
    model_data = model.getData()
    for line in model_data:
        if "房号" in model_data[line].keys():
            excel_data[line] = model_data[line]
        else:
            break
    print(excel_data)

    # 房号 单位名称 房租 物管费 合同
    month = int(month_combox.get())
    pay_way = pay_way_combox.get()
    query_result_save = {}
    for line in excel_data:
        query_result_save[line] = {}
        query_result_save[line]["房号"] = excel_data[line]["房号"]
        query_result_save[line]["公司名称"] = excel_data[line]["公司名称"]
        query_result_save[line]["房租"] = excel_data[line]["房租"]
        query_result_save[line]["物管费"] = excel_data[line]["物管费"]
        contract_date = datetime.datetime.strptime(
            excel_data[line]["合同日期"], '%Y-%m-%d')
        if pay_way == "季度":
            query_result_save[line]["房租缴费起止日期"] = get_begin_end(
                contract_date, month, 1)
        elif pay_way == "半年":
            query_result_save[line]["房租缴费起止日期"] = get_begin_end(
                contract_date, month, 2)
        elif pay_way == "年度":
            query_result_save[line]["房租缴费起止日期"] = get_begin_end(
                contract_date, month, 3)
        elif pay_way == "所有":
            if excel_data[line]["付款方式"] == "季度":
                query_result_save[line]["房租缴费起止日期"] = get_begin_end(
                    contract_date, month, 1)
            if excel_data[line]["付款方式"] == "半年":
                query_result_save[line]["房租缴费起止日期"] = get_begin_end(
                    contract_date, month, 2)
            if excel_data[line]["付款方式"] == "年度":
                query_result_save[line]["房租缴费起止日期"] = get_begin_end(
                    contract_date, month, 3)
        print(query_result_save[line])
    excel_df = pd.DataFrame(query_result_save).transpose()
    if file_path:
        print(file_path)
        try:
            excel_df.to_excel(file_path, index=False)
        except:
            tkinter.messagebox.showerror("错误", "文件已被打开，未被保存，请重新查询")

        neat_file(file_path)


def save(model: TableModel):
    global dict_variable
    print("saving")
    file_path = filedialog.asksaveasfilename(defaultextension='.xlsx',
                                             filetypes=[('Excel文件', '*.xlsx'), ('所有文件', '*.*')])
    excel_data = {}
    model_data = model.getData()
    for line in model_data:
        if "房号" in model_data[line].keys():
            excel_data[line] = model_data[line]
        else:
            break
    print(excel_data)
    excel_df = pd.DataFrame(excel_data).transpose()
    if file_path:
        print(file_path)
        excel_df.to_excel(file_path, index=False)


def add_new_row(table: TableCanvas):
    table.addRow()


def delete_current_row(table: TableCanvas):
    table.deleteRow()
    table.redrawTable()
